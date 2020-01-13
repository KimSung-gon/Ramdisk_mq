import openpyxl

global row
global col

select = int(input("select 1-ramdisk_mq or 2-null_blk : "))
select2 = int(input("select 1-libaio or 2-psync : "))
select3 = int(input("select 1-read or 2-write : " ))

if select == 1:
    module = "ramdisk_mq"
    device = "ramdisk_mq"
elif select == 2:
    module = "null_blk"
    device = "nullb0"

if select2 == 1:
    ioengine = "libaio"
elif select2 == 2:
    ioengine = "psync"

def current():
    return  col + str(row)

def nextRow():
    global row
    row = str(int(row) + 1)

def nextCol():
    global col
    col = chr(ord(col) + 1)

def initCol():
    global col
    col = "A"

def initRow():
    global row
    row = str(int(row) - 24)

wb = openpyxl.Workbook()
ws = wb.active

row = "1"
col = "A"

nr_hw_queue = [120,60,40,30,24,20,15,12,10,8,7,6,5,4,3,2]
nr_requests = [65536,32768,16384,8192,4096,2048,1024,512,256,128,64,32,16,8,4,2,1]
nr_cores = [5,10,15,20,25,30,35,40,45,50,55,60,65,70,75,80,85,90,95,100,105,110,115,120]

nextRow()
nextRow()

base_dir = "/media/sda/kesl/re_test_20181206/bio/"
base_dir += (module + "_" + ioengine)

if select3 == 2:
    base_dir += "_write"

def write_commands(nr_hw_queues, device, module):
    ws[current()] = "mq_" + str(nr_hw_queues) + "nr_hw_queue_raw"
    nextRow()
    if select == 1:
        ws[current()] = "modprobe ramdisk_mq nr_hw_queue=" + str(nr_hw_queues) + " queue_mode=2 ramdisk_mq_size=419430400 nr_hw_queue_depth=x"
    elif select == 2:
        ws[current()] = "modprobe null_blk completion_nsec=0 irqmode=0 nr_devices=1 bs=4096 queue_mode=2 submit_queues=" + str(nr_hw_queues) + " hw_queue_depth=x"
    nextRow()
    ws[current()] = "sudo fio --output=xxxx --name=myjob --filename=/dev/" + device + " --ioengine==" + ioengine + " --direct=1 norandommap --randrepeat=0 --runtime=10 --time_based --blocksize=4K --rw=randread --iodepth=128 --numjobs=x"

def write_cores():
    for temp in nr_cores:
        ws[current()] = temp
        nextRow()

def write_nr_requests():
    for temp in nr_requests:
        ws[current()] = temp
        nextCol()

def write_value_per_core(temp_dir):
    temp_dir_core = temp_dir + "/job:_" + str(z) + "_dp:_128.txt"
    f = open(temp_dir_core, "r")
    lines = f.readlines()
    iops_line = lines[6]
    start = iops_line.find('iops=')
    end = iops_line.find(',', start)
    iops = iops_line[(start+5):(end-1)]
    if iops_line[end-1] == 'k':
        iops = str(round(float(iops)/1000,1))
    ws[current()] = iops
    f.close()

for x in nr_hw_queue:

    write_commands(x, device, module)

    nextRow()
    nextCol()

    write_nr_requests()

    initCol()
    nextRow()

    temp_row = row
    temp_col = col

    write_cores()

    row = temp_row
    col = temp_col
    nextCol()

    for y in nr_requests:

        temp_dir = base_dir + "/hw_queues_" + str(x) + "_queue_depth_" + str(y)

        for z in nr_cores:
            write_value_per_core(temp_dir)
            nextRow()

        initRow()
        nextCol()
    for i in range(24):
        nextRow()
    initCol()
    nextRow()

wb.save("./new_test_file.xlsx")

